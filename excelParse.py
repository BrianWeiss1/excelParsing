import openpyxl
def getCyberScores(excelFile: str='round1.xlsx', preCiscoScore: float = 184, postCiscoScore: float = 193.33):
    path = f"{excelFile}"
    
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    finalScores = []
    divsionScores = []
    newEnglandScores = []
    StateScores = []
    for i in range(1, 20000):
        cell_obj = sheet_obj.cell(row=7+i, column=1)
        # print(cell_obj.content)
        # print(type(str(cell_obj)))
        # print(str(cell_obj))
        if str(cell_obj.value) == '16-0028':
            print(i)
            for j in range(6):
                coll_obj = sheet_obj.cell(row=7+i, column=7+j)
                print(coll_obj.value)
                # print(round(float(sheet_obj.cell(row=7+i, column=3).value)+float(sheet_obj.cell(row=7+i, column=4).value), 5))
        try:
            final_score = float(sheet_obj.cell(row=7+i, column=6).value)

        except:
            final_score = 0
        division = sheet_obj.cell(row=7+i, column=3)
        location = sheet_obj.cell(row=7+i, column=2)
        tier = sheet_obj.cell(row=7+i, column=7)
        if tier.value == 'Gold':
            try:
                finalScores.append(float(final_score))
                if division.value == 'Open':
                    divsionScores.append(float(final_score))
                if location.value == 'CT':
                    StateScores.append(float(final_score))
                if location.value == 'CT' or location.value == 'ME' or location.value == 'MA' or location.value == 'NH' or location.value == 'RI' or location.value == 'VT':
                    newEnglandScores.append(float(final_score))
            except (TypeError, ValueError):
                pass 



    finalScores.sort(reverse=True)
    for i in range(len(finalScores)):
        if finalScores[i] == postCiscoScore:
            specialI = i

    print("Post-Cisco")
    print("All: " + str(specialI) +  "/" + str(len(finalScores)))

    divsionScores.sort(reverse=True)
    for i in range(len(divsionScores)):
        if divsionScores[i] == postCiscoScore:
            specialI = i

    print("Open: " + str(specialI) +  "/" + str(len(divsionScores)))


    StateScores.sort(reverse=True)
    for i in range(len(StateScores)):
        if StateScores[i] == postCiscoScore:
            specialI = i

    print("CT: " + str(specialI) +  "/" + str(len(StateScores)))

    newEnglandScores.sort(reverse=True)
    for i in range(len(newEnglandScores)):
        if newEnglandScores[i] == postCiscoScore:
            specialI = i

    print("New England: " + str(specialI) +  "/" + str(len(newEnglandScores)))


    # Pre-Cisco

    sheet_obj = wb_obj.active
    finalScores = []
    divsionScores = []
    newEnglandScores = []
    StateScores = []
    for i in range(1, 10000):
        cell_obj = sheet_obj.cell(row=7+i, column=1)
        # print(cell_obj.content)
        # print(type(str(cell_obj)))
        # print(str(cell_obj))
        if str(cell_obj.value) == '16-0028':
            print(i)
            for j in range(6):
                coll_obj = sheet_obj.cell(row=7+i, column=1+j)
                print(coll_obj.value)
        final_score = sheet_obj.cell(row=7+i, column=4)
        division = sheet_obj.cell(row=7+i, column=3)
        location = sheet_obj.cell(row=7+i, column=2)
        tier = sheet_obj.cell(row=7+i, column=7)
        if tier.value == 'Gold':
            try:
                finalScores.append(float(final_score.value))
                if division.value == 'Open':
                    divsionScores.append(float(final_score.value))
                if location.value == 'CT':
                    StateScores.append(float(final_score.value))
                if location.value == 'CT' or location.value == 'ME' or location.value == 'MA' or location.value == 'NH' or location.value == 'RI' or location.value == 'VT':
                    newEnglandScores.append(float(final_score.value))
            except (TypeError, ValueError):
                pass 


    finalScores.sort(reverse=True)
    for i in range(len(finalScores)):
        if finalScores[i] == preCiscoScore:
            specialI = i

    print("Pre-Cisco: ")
    print("All: " + str(specialI) +  "/" + str(len(finalScores)))

    divsionScores.sort(reverse=True)
    for i in range(len(divsionScores)):
        if divsionScores[i] == preCiscoScore:
            specialI = i

    print("Open: " + str(specialI) +  "/" + str(len(divsionScores)))


    StateScores.sort(reverse=True)
    for i in range(len(StateScores)):
        if StateScores[i] == preCiscoScore:
            specialI = i

    print("CT: " + str(specialI) +  "/" + str(len(StateScores)))

    newEnglandScores.sort(reverse=True)
    for i in range(len(newEnglandScores)):
        if newEnglandScores[i] == preCiscoScore:
            specialI = i

    print("New England: " + str(specialI) +  "/" + str(len(newEnglandScores)))

# Round1 = getCyberScores('round1.xlsx', 184, 193.33)
# Round2 = getCyberScores('round2.xlsx', 149, 155)
Round3 = getCyberScores('stateRound.xlsx', 99, 114)